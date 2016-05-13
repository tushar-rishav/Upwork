# -*- coding: utf-8 -*-
"""
Extract textual data from a given html document
sample url:
    https://www.sec.gov/Archives/edgar/data/1594109/000156459015001303/grub-10k_20141231.htm

"""

from docx import Document
import openpyxl
from openpyxl import Workbook
from openpyxl.writer.write_only import WriteOnlyCell
import os
import pyautogui
from selenium import webdriver
import sys
from tkinter import Tk
from time import sleep

driver = webdriver.Firefox()
driver.implicitly_wait(40)


def browse(url):
    """
    Browser handler
    :param url: A url to load in browser.
    """
    try:
        driver.get(url)
    except Exception as e:
        print(e)

def load_url_from_excel(excel_file="data.xlsx"):
    """
    Extract URLs from a given excel file.
    :param excel_file: The excel file that contains URLs.
    :return: A list of URLs.
    """
    if os.path.exists(excel_file) is not None:
        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.get_sheet_by_name('Sheet1')
        return [sheet['A{}'.format(i)].value.encode('utf-8') for i in range(2,12)]

    print("{} does not exists.".format(excel_file))
    sys.exit(1)

def read_clipboard():
    itk = Tk()
    itk.withdraw()
    try:
        while not itk.selection_get(selection="CLIPBOARD"):
            sleep(0.1)
    except Exception as e:
        print(e)

    result = itk.selection_get(selection="CLIPBOARD")
    itk.destroy()
    return result

def save_doc_and_excel(filename):
    """
    Create a new word doc with the given filename.
    :param filename: The name of new doc.
    """
    # Save doc
    document = Document()
    content = read_clipboard()
    document.add_paragraph(content)
    document.save(filename + '.doc')

    # Save excel sheets
    wb = Workbook(write_only=True)
    ws = wb.create_sheet()
    cell = WriteOnlyCell(ws, value=content)
    ws.append([cell])
    wb.save(filename + '.xlsx')

def main():
    width, height = pyautogui.size()
    URLS = load_url_from_excel()
    i = 0
    pyautogui.PAUSE = 1
    for url in URLS:
        browse(url)
        pyautogui.click(x=width/3, y=height/3)
        pyautogui.hotkey('ctrl', 'a')    # Select all text
        pyautogui.hotkey('ctrl', 'c')    # Copy all text
        sleep(1)
        print("Saving output_{}.doc/xlsx".format(i))
        save_doc_and_excel("output_{}".format(i))
        i += 1
    print("Closing browser")
    driver.close()
    print("Browser closed")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("Closed")